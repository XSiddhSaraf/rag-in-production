"""Excel file generation for analysis results."""

from pathlib import Path
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from models import ProjectAnalysis, EvaluationMetrics, LLMJudgeResult, Risk
from loguru import logger
from datetime import datetime


class ExcelGenerator:
    """Generate Excel reports from analysis results."""
    
    def __init__(self):
        """Initialize Excel generator."""
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True, size=12)
        self.high_risk_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        self.low_risk_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def _format_header_row(self, ws, row_num: int, headers: List[str]):
        """Format header row with styling."""
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border
    
    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 60)  # Cap at 60
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def generate_excel(
        self,
        analysis: ProjectAnalysis,
        evaluation_metrics: EvaluationMetrics = None,
        llm_judge: LLMJudgeResult = None,
        output_path: Path = None
    ) -> Path:
        """Generate Excel file with analysis results.
        
        Args:
            analysis: Project analysis results
            evaluation_metrics: RAG evaluation metrics
            llm_judge: LLM-as-judge results
            output_path: Output file path
            
        Returns:
            Path to generated Excel file
        """
        logger.info("Generating Excel report...")
        
        wb = Workbook()
        
        # Sheet 1: Summary
        ws_summary = wb.active
        ws_summary.title = "Summary"
        self._create_summary_sheet(ws_summary, analysis)
        
        # Sheet 2: High Risks
        ws_high_risks = wb.create_sheet("High Risks")
        self._create_risks_sheet(ws_high_risks, analysis.high_risks, "HIGH")
        
        # Sheet 3: Low Risks
        ws_low_risks = wb.create_sheet("Low Risks")
        self._create_risks_sheet(ws_low_risks, analysis.low_risks, "LOW")
        
        # Sheet 4: Evaluation Metrics (if available)
        if evaluation_metrics or llm_judge:
            ws_eval = wb.create_sheet("Evaluation")
            self._create_evaluation_sheet(ws_eval, evaluation_metrics, llm_judge)
        
        # Determine output path
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{analysis.project_name.replace(' ', '_')}_{timestamp}.xlsx"
            output_path = Path("outputs") / filename
        
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Save
        wb.save(output_path)
        logger.info(f"Excel report saved: {output_path}")
        
        return output_path
    
    def _create_summary_sheet(self, ws, analysis: ProjectAnalysis):
        """Create summary sheet."""
        # Headers
        headers = ["Project Name", "Description", "Contains AI", "High Risks", "Low Risks"]
        self._format_header_row(ws, 1, headers)
        
        # Data row
        row = 2
        ws.cell(row, 1, analysis.project_name).border = self.border
        ws.cell(row, 2, analysis.description).border = self.border
        ws.cell(row, 3, "YES" if analysis.contains_ai else "NO").border = self.border
        ws.cell(row, 4, len(analysis.high_risks)).border = self.border
        ws.cell(row, 5, len(analysis.low_risks)).border = self.border
        
        # Highlight Contains AI
        ai_cell = ws.cell(row, 3)
        if analysis.contains_ai:
            ai_cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
            ai_cell.font = Font(bold=True)
        
        # AI Confidence
        ws.cell(row + 2, 1, "AI Detection Confidence:").font = Font(bold=True)
        ws.cell(row + 2, 2, f"{analysis.ai_confidence:.1%}")
        
        self._auto_adjust_columns(ws)
    
    def _create_risks_sheet(self, ws, risks: List[Risk], risk_level: str):
        """Create risks detail sheet."""
        # Headers
        headers = ["#", "Risk Description", "Category", "EU AI Act Reference", "Confidence Score"]
        self._format_header_row(ws, 1, headers)
        
        # Fill color based on risk level
        fill = self.high_risk_fill if risk_level == "HIGH" else self.low_risk_fill
        
        # Data rows
        for idx, risk in enumerate(risks, start=1):
            row = idx + 1
            
            ws.cell(row, 1, idx).border = self.border
            ws.cell(row, 2, risk.description).border = self.border
            ws.cell(row, 3, risk.category).border = self.border
            ws.cell(row, 4, risk.eu_act_reference or "N/A").border = self.border
            
            confidence_cell = ws.cell(row, 5, f"{risk.confidence_score:.1%}" if risk.confidence_score else "N/A")
            confidence_cell.border = self.border
            
            # Apply risk-level fill to description
            ws.cell(row, 2).fill = fill
        
        # Add summary
        summary_row = len(risks) + 3
        ws.cell(summary_row, 1, f"Total {risk_level} Risks:").font = Font(bold=True)
        ws.cell(summary_row, 2, len(risks)).font = Font(bold=True, size=14)
        
        self._auto_adjust_columns(ws)
    
    def _create_evaluation_sheet(self, ws, metrics: EvaluationMetrics, llm_judge: LLMJudgeResult):
        """Create evaluation metrics sheet."""
        row = 1
        
        # RAG Metrics section
        if metrics:
            ws.cell(row, 1, "RAG Evaluation Metrics").font = Font(bold=True, size=14)
            row += 1
            
            self._format_header_row(ws, row, ["Metric", "Score"])
            row += 1
            
            metrics_dict = metrics.model_dump()
            for metric_name, score in metrics_dict.items():
                if score is not None:
                    ws.cell(row, 1, metric_name.replace('_', ' ').title()).border = self.border
                    score_cell = ws.cell(row, 2, f"{score:.2%}")
                    score_cell.border = self.border
                    
                    # Color code
                    if score >= 0.8:
                        score_cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
                    elif score >= 0.6:
                        score_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    else:
                        score_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    
                    row += 1
            
            row += 2
        
        # LLM Judge section
        if llm_judge:
            ws.cell(row, 1, "LLM-as-a-Judge Evaluation").font = Font(bold=True, size=14)
            row += 1
            
            self._format_header_row(ws, row, ["Criterion", "Score"])
            row += 1
            
            judge_dict = {
                "Accuracy": llm_judge.accuracy_score,
                "Completeness": llm_judge.completeness_score,
                "Consistency": llm_judge.consistency_score,
                "Overall": llm_judge.overall_score
            }
            
            for criterion, score in judge_dict.items():
                ws.cell(row, 1, criterion).border = self.border
                score_cell = ws.cell(row, 2, f"{score:.2%}")
                score_cell.border = self.border
                
                if score >= 0.8:
                    score_cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
                elif score >= 0.6:
                    score_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:
                    score_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                row += 1
            
            row += 2
            ws.cell(row, 1, "Reasoning:").font = Font(bold=True)
            row += 1
            reasoning_cell = ws.cell(row, 1, llm_judge.reasoning)
            reasoning_cell.alignment = Alignment(wrap_text=True, vertical='top')
            ws.merge_cells(f'A{row}:B{row + 2}')
        
        self._auto_adjust_columns(ws)


# Global Excel generator instance
excel_generator = ExcelGenerator()
